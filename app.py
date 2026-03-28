"""
豆瓣电影数据分析与推荐系统 - Web版本
使用 PyTorch CNN+LSTM 情感分析模型 + 神经网络推荐引擎
"""

from flask import Flask, render_template, request, jsonify, session
from werkzeug.utils import secure_filename
import matplotlib.pyplot as plt
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
from io import BytesIO
import base64
import numpy as np
from datetime import datetime
import random
from collections import Counter
import json
import os
import re
import jieba
import pickle
import time
import pandas as pd

# PyTorch
import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import Dataset, DataLoader
from torch.nn.utils.rnn import pad_sequence

# 爬虫相关库
try:
    import requests
    from bs4 import BeautifulSoup
    import pandas as pd
    REQUESTS_AVAILABLE = True
    PANDAS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False
    PANDAS_AVAILABLE = False

# 设置中文字体
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False

# 创建Flask应用
app = Flask(__name__)
app.secret_key = 'douban_movie_system_secret_key_2024'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# 确保上传目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# ==================== 豆瓣真实风格评论数据 ====================
DOUBAN_REAL_REVIEWS = {
    "positive": [
        "太震撼了！看完久久不能平静，绝对是年度最佳！",
        "经典中的经典，每次看都有新的感悟，五星推荐！",
        "导演的功力太强了，每一个镜头都是艺术品，演员的表演也非常到位。",
        "哭了整整两个小时，这部电影真的太打动人心了。",
        "神作！没有任何一部电影能给我这样的观影体验。",
        "剧本太牛了，情节环环相扣，结局出人意料又情理之中。",
        "看完只想说：这才是电影！国产电影的骄傲！",
        "满分好评，已经二刷了，每次看都有新发现。",
        "太喜欢了！从配乐到画面都无可挑剔。",
        "这部电影改变了我的人生观，强烈推荐每个人都去看！"
    ],
    "neutral": [
        "还行吧，中规中矩，没有想象中那么好。",
        "一般般，看完就忘，没有什么印象深刻的点。",
        "普普通通，可以一看，但不会二刷。",
        "期望太高了，实际看下来有点失望。",
        "不算差但也不算好，打发时间还可以。",
        "剧情有点老套，没什么新意。",
        "演员演技还行，但剧本太弱了。",
        "特效不错，但剧情拖沓，节奏太慢。",
        "看完没什么感觉，可能不是我喜欢的类型。",
        "及格线以上的作品，但谈不上经典。"
    ],
    "negative": [
        "太失望了！完全浪费时间和金钱。",
        "什么烂片？剧情逻辑完全不通，看得我尴尬癌都犯了。",
        "演员演技太尬了，完全看不下去。",
        "一分都不想给，浪费时间！",
        "剧本太差了，全是套路，毫无新意。",
        "特效五毛，剧情狗血，烂片中的战斗机。",
        "导演想表达什么完全看不懂，故弄玄虚。",
        "太无聊了，看了半小时就想走。",
        "改编得一塌糊涂，完全毁了原著。",
        "全程尬演，尴尬到脚趾抠地。"
    ]
}

USER_NAMES = [
    "豆瓣用户", "影迷小张", "电影爱好者", "观影达人", "评分机器人", "影评人小王",
    "路人甲", "电影发烧友", "爱看电影的猫", "深夜观影者", "周末影迷", "电影控"
]

# ==================== 扩展本地电影数据库 ====================
LOCAL_MOVIE_CACHE = {
    "肖申克的救赎": {"id": "1292052", "title": "肖申克的救赎", "year": "1994", "rating": 9.7, "director": "弗兰克·德拉邦特", "actors": ["蒂姆·罗宾斯", "摩根·弗里曼"], "genre": "剧情", "country": "美国", "language": "英语", "duration": "142分钟", "description": "希望让人自由，一部关于希望与救赎的经典之作。"},
    "霸王别姬": {"id": "1291546", "title": "霸王别姬", "year": "1993", "rating": 9.6, "director": "陈凯歌", "actors": ["张国荣", "张丰毅", "巩俐"], "genre": "剧情", "country": "中国", "language": "汉语普通话", "duration": "171分钟", "description": "风华绝代，一曲人生悲歌。"},
    "阿甘正传": {"id": "1292720", "title": "阿甘正传", "year": "1994", "rating": 9.5, "director": "罗伯特·泽米吉斯", "actors": ["汤姆·汉克斯"], "genre": "剧情", "country": "美国", "language": "英语", "duration": "142分钟", "description": "人生就像一盒巧克力，你永远不知道下一块是什么味道。"},
    "这个杀手不太冷": {"id": "1295644", "title": "这个杀手不太冷", "year": "1994", "rating": 9.4, "director": "吕克·贝松", "actors": ["让·雷诺", "娜塔莉·波特曼"], "genre": "剧情/动作", "country": "法国", "language": "英语", "duration": "110分钟", "description": "杀手与小女孩的温情故事。"},
    "泰坦尼克号": {"id": "1292722", "title": "泰坦尼克号", "year": "1997", "rating": 9.5, "director": "詹姆斯·卡梅隆", "actors": ["莱昂纳多", "凯特·温丝莱特"], "genre": "爱情/灾难", "country": "美国", "language": "英语", "duration": "194分钟", "description": "永恒的爱情，永不沉没的经典。"},
    "盗梦空间": {"id": "3541415", "title": "盗梦空间", "year": "2010", "rating": 9.4, "director": "克里斯托弗·诺兰", "actors": ["莱昂纳多"], "genre": "科幻/悬疑", "country": "美国", "language": "英语", "duration": "148分钟", "description": "梦境与现实，一场关于意识的大胆探索。"},
    "楚门的世界": {"id": "1292064", "title": "楚门的世界", "year": "1998", "rating": 9.3, "director": "彼得·威尔", "actors": ["金·凯瑞"], "genre": "剧情/科幻", "country": "美国", "language": "英语", "duration": "103分钟", "description": "真实与虚假的边界在哪里？"},
    "千与千寻": {"id": "1291561", "title": "千与千寻", "year": "2001", "rating": 9.4, "director": "宫崎骏", "actors": ["柊瑠美"], "genre": "动画/奇幻", "country": "日本", "language": "日语", "duration": "125分钟", "description": "成长的旅程，宫崎骏的巅峰之作。"},
    "星际穿越": {"id": "1889243", "title": "星际穿越", "year": "2014", "rating": 9.4, "director": "克里斯托弗·诺兰", "actors": ["马修·麦康纳"], "genre": "科幻", "country": "美国", "language": "英语", "duration": "169分钟", "description": "穿越时空的爱，人类命运的终极探索。"},
    "海上钢琴师": {"id": "1292001", "title": "海上钢琴师", "year": "1998", "rating": 9.3, "director": "朱塞佩·托纳多雷", "actors": ["蒂姆·罗斯"], "genre": "剧情/音乐", "country": "意大利", "language": "英语", "duration": "125分钟", "description": "1900的故事，音乐与自由的传奇。"},
    "让子弹飞": {"id": "3742360", "title": "让子弹飞", "year": "2010", "rating": 9.0, "director": "姜文", "actors": ["姜文", "葛优", "周润发"], "genre": "剧情/喜剧", "country": "中国", "language": "汉语普通话", "duration": "132分钟", "description": "站着把钱挣了！"},
    "流浪地球": {"id": "26266893", "title": "流浪地球", "year": "2019", "rating": 7.9, "director": "郭帆", "actors": ["吴京", "屈楚萧", "李光洁"], "genre": "科幻", "country": "中国", "language": "汉语普通话", "duration": "125分钟", "description": "带着地球去流浪，中国科幻的里程碑。"},
    "哪吒之魔童降世": {"id": "26794435", "title": "哪吒之魔童降世", "year": "2019", "rating": 8.4, "director": "饺子", "actors": ["吕艳婷", "囧森瑟夫"], "genre": "动画/奇幻", "country": "中国", "language": "汉语普通话", "duration": "110分钟", "description": "我命由我不由天！"},
    "我不是药神": {"id": "26752088", "title": "我不是药神", "year": "2018", "rating": 9.0, "director": "文牧野", "actors": ["徐峥", "王传君", "周一围"], "genre": "剧情", "country": "中国", "language": "汉语普通话", "duration": "117分钟", "description": "现实题材的震撼之作。"},
    "绿皮书": {"id": "27060077", "title": "绿皮书", "year": "2018", "rating": 8.9, "director": "彼得·法拉利", "actors": ["维果·莫腾森", "马赫沙拉·阿里"], "genre": "剧情/喜剧", "country": "美国", "language": "英语", "duration": "130分钟", "description": "跨越种族与阶层的友谊。"},
    "三傻大闹宝莱坞": {"id": "3793023", "title": "三傻大闹宝莱坞", "year": "2009", "rating": 9.2, "director": "拉吉库马尔·希拉尼", "actors": ["阿米尔·汗", "马德哈万"], "genre": "剧情/喜剧", "country": "印度", "language": "印地语", "duration": "171分钟", "description": "追求卓越，成功会不请自来。"},
    "放牛班的春天": {"id": "1291548", "title": "放牛班的春天", "year": "2004", "rating": 9.3, "director": "克里斯托夫·巴拉蒂", "actors": ["热拉尔·朱尼奥", "让-巴蒂斯特·莫尼耶"], "genre": "剧情/音乐", "country": "法国", "language": "法语", "duration": "97分钟", "description": "音乐治愈心灵。"},
    "忠犬八公的故事": {"id": "3011091", "title": "忠犬八公的故事", "year": "2009", "rating": 9.4, "director": "莱塞·霍尔斯道姆", "actors": ["理查·基尔", "琼·艾伦"], "genre": "剧情", "country": "美国", "language": "英语", "duration": "93分钟", "description": "等待是最长情的告白。"},
    "美丽人生": {"id": "1292063", "title": "美丽人生", "year": "1997", "rating": 9.5, "director": "罗伯托·贝尼尼", "actors": ["罗伯托·贝尼尼", "尼可莱塔·布拉斯基"], "genre": "剧情/喜剧", "country": "意大利", "language": "意大利语", "duration": "116分钟", "description": "即使在黑暗中，也要相信美好。"},
    "怦然心动": {"id": "3319755", "title": "怦然心动", "year": "2010", "rating": 9.1, "director": "罗伯·莱纳", "actors": ["玛德琳·卡罗尔", "卡兰·麦克奥利菲"], "genre": "爱情/喜剧", "country": "美国", "language": "英语", "duration": "90分钟", "description": "初恋的美好，成长的喜悦。"},
}

# 模糊匹配辅助函数
def fuzzy_match_movie(movie_name):
    from difflib import SequenceMatcher
    movie_name_lower = movie_name.lower()
    best_match = None
    best_score = 0
    
    for cached_name in LOCAL_MOVIE_CACHE.keys():
        score = SequenceMatcher(None, movie_name_lower, cached_name.lower()).ratio()
        if cached_name.lower() in movie_name_lower or movie_name_lower in cached_name.lower():
            score = max(score, 0.7)
        if score > best_score and score > 0.5:
            best_score = score
            best_match = cached_name
    
    return best_match, best_score


# ==================== 神经网络推荐引擎 ====================

class NeuralRecommender(nn.Module):
    """基于神经网络的推荐模型 - 矩阵分解 + MLP"""
    
    def __init__(self, num_users, num_items, embedding_dim=64, hidden_dims=[128, 64]):
        super(NeuralRecommender, self).__init__()
        
        # 用户和物品的嵌入层
        self.user_embedding = nn.Embedding(num_users, embedding_dim)
        self.item_embedding = nn.Embedding(num_items, embedding_dim)
        
        # 特征嵌入层（用于物品特征，如类型、导演等）
        self.genre_embedding = nn.Embedding(15, 8)  # 15种类型
        self.director_embedding = nn.Embedding(30, 8)  # 30位导演
        self.year_embedding = nn.Embedding(50, 8)  # 50个年份
        
        # MLP层
        input_dim = embedding_dim * 2 + 8 + 8 + 8  # 用户嵌入 + 物品嵌入 + 类型 + 导演 + 年份
        layers = []
        for hidden_dim in hidden_dims:
            layers.append(nn.Linear(input_dim, hidden_dim))
            layers.append(nn.ReLU())
            layers.append(nn.Dropout(0.3))
            input_dim = hidden_dim
        
        layers.append(nn.Linear(input_dim, 1))
        layers.append(nn.Sigmoid())
        
        self.mlp = nn.Sequential(*layers)
        
    def forward(self, user_ids, item_ids, genre_ids, director_ids, year_ids):
        # 嵌入
        user_vec = self.user_embedding(user_ids)
        item_vec = self.item_embedding(item_ids)
        genre_vec = self.genre_embedding(genre_ids)
        director_vec = self.director_embedding(director_ids)
        year_vec = self.year_embedding(year_ids)
        
        # 拼接所有特征
        concat_vec = torch.cat([user_vec, item_vec, genre_vec, director_vec, year_vec], dim=1)
        
        # MLP预测
        pred = self.mlp(concat_vec)
        return pred.squeeze()


class HybridRecommender:
    """混合推荐系统 - 协同过滤 + 基于内容 + 神经网络"""
    
    def __init__(self):
        self.user_embedding = None
        self.item_embedding = None
        self.model = None
        self.is_trained = False
        self.device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
        
        # 类型映射
        self.genre_map = {
            "剧情": 0, "动作": 1, "爱情": 2, "科幻": 3, "悬疑": 4,
            "动画": 5, "奇幻": 6, "喜剧": 7, "灾难": 8, "音乐": 9,
            "历史": 10, "犯罪": 11, "惊悚": 12, "冒险": 13, "家庭": 14
        }
        
        # 导演映射
        self.director_map = {}
        self.year_map = {}
        
    def _build_mappings(self, movies, users):
        """构建特征映射"""
        # 导演映射
        directors = set()
        for movie in movies:
            directors.add(movie["director"])
        for i, director in enumerate(sorted(directors)):
            self.director_map[director] = i
        
        # 年份映射
        years = set()
        for movie in movies:
            years.add(movie["year"])
        for i, year in enumerate(sorted(years)):
            self.year_map[year] = i
    
    def _get_genre_id(self, genre_str):
        """获取类型ID"""
        for genre in self.genre_map:
            if genre in genre_str:
                return self.genre_map[genre]
        return 0
    
    def train(self, user_interactions, movies, epochs=50, embedding_dim=64):
        """训练神经网络推荐模型"""
        print("开始训练神经网络推荐模型...")
        
        # 构建映射
        self._build_mappings(movies, user_interactions)
        
        # 准备训练数据
        user_ids = []
        item_ids = []
        ratings = []
        genre_ids = []
        director_ids = []
        year_ids = []
        
        # 构建用户和物品ID映射
        user_map = {}
        item_map = {}
        
        for user_id, interactions in user_interactions.items():
            if user_id not in user_map:
                user_map[user_id] = len(user_map)
            
            for item_id, rating in interactions.items():
                if item_id not in item_map:
                    item_map[item_id] = len(item_map)
                
                # 获取电影特征
                movie = next((m for m in movies if m["id"] == item_id), None)
                if movie:
                    user_ids.append(user_map[user_id])
                    item_ids.append(item_map[item_id])
                    ratings.append(rating / 10.0)  # 归一化到0-1
                    genre_ids.append(self._get_genre_id(movie["genre"]))
                    director_ids.append(self.director_map.get(movie["director"], 0))
                    year_ids.append(self.year_map.get(movie["year"], 0))
        
        if len(user_ids) < 10:
            print("训练数据不足，使用降级推荐")
            return False
        
        # 转换为Tensor
        user_ids_tensor = torch.tensor(user_ids, dtype=torch.long).to(self.device)
        item_ids_tensor = torch.tensor(item_ids, dtype=torch.long).to(self.device)
        ratings_tensor = torch.tensor(ratings, dtype=torch.float32).to(self.device)
        genre_ids_tensor = torch.tensor(genre_ids, dtype=torch.long).to(self.device)
        director_ids_tensor = torch.tensor(director_ids, dtype=torch.long).to(self.device)
        year_ids_tensor = torch.tensor(year_ids, dtype=torch.long).to(self.device)
        
        # 创建模型
        self.model = NeuralRecommender(
            num_users=len(user_map),
            num_items=len(item_map),
            embedding_dim=embedding_dim
        ).to(self.device)
        
        # 训练
        optimizer = optim.Adam(self.model.parameters(), lr=0.01)
        criterion = nn.MSELoss()
        
        for epoch in range(epochs):
            self.model.train()
            optimizer.zero_grad()
            
            pred = self.model(user_ids_tensor, item_ids_tensor, 
                             genre_ids_tensor, director_ids_tensor, year_ids_tensor)
            loss = criterion(pred, ratings_tensor)
            loss.backward()
            optimizer.step()
            
            if (epoch + 1) % 10 == 0:
                print(f"Epoch {epoch+1}/{epochs}, Loss: {loss.item():.4f}")
        
        self.is_trained = True
        self.user_map = user_map
        self.item_map = item_map
        self.reverse_item_map = {v: k for k, v in item_map.items()}
        
        print("神经网络推荐模型训练完成")
        return True
    
    def predict_rating(self, user_id, movie, user_ratings):
        """预测用户对电影的评分"""
        if not self.is_trained or self.model is None:
            # 降级：基于内容相似度
            return self._content_based_score(user_id, movie, user_ratings)
        
        try:
            if user_id not in self.user_map:
                return self._content_based_score(user_id, movie, user_ratings)
            
            if movie["id"] not in self.item_map:
                return self._content_based_score(user_id, movie, user_ratings)
            
            self.model.eval()
            with torch.no_grad():
                user_tensor = torch.tensor([self.user_map[user_id]], dtype=torch.long).to(self.device)
                item_tensor = torch.tensor([self.item_map[movie["id"]]], dtype=torch.long).to(self.device)
                genre_tensor = torch.tensor([self._get_genre_id(movie["genre"])], dtype=torch.long).to(self.device)
                director_tensor = torch.tensor([self.director_map.get(movie["director"], 0)], dtype=torch.long).to(self.device)
                year_tensor = torch.tensor([self.year_map.get(movie["year"], 0)], dtype=torch.long).to(self.device)
                
                pred = self.model(user_tensor, item_tensor, genre_tensor, director_tensor, year_tensor)
                score = pred.item() * 10
                
                # 结合内容相似度
                content_score = self._content_based_score(user_id, movie, user_ratings)
                final_score = score * 0.7 + content_score * 0.3
                
                return final_score
        except Exception as e:
            print(f"神经网络预测失败: {e}")
            return self._content_based_score(user_id, movie, user_ratings)
    
    def _content_based_score(self, user_id, movie, user_ratings):
        """基于内容的相似度评分"""
        if not user_ratings:
            return movie["rating"] * 0.8
        
        # 获取用户喜欢的电影特征
        liked_movies = []
        for item_id, rating in user_ratings.items():
            if rating >= 7:
                movie_obj = next((m for m in LOCAL_MOVIE_CACHE.values() if m.get("id") == str(item_id)), None)
                if movie_obj:
                    liked_movies.append(movie_obj)
        
        if not liked_movies:
            return movie["rating"] * 0.8
        
        # 计算相似度
        similarity = 0
        for liked in liked_movies:
            # 类型相似度
            genre_sim = 1 if liked["genre"].split('/')[0] == movie["genre"].split('/')[0] else 0.3
            # 导演相似度
            director_sim = 1 if liked["director"] == movie["director"] else 0.2
            # 年份相似度
            year_diff = abs(int(liked["year"]) - int(movie["year"]))
            year_sim = max(0, 1 - year_diff / 50)
            
            sim = genre_sim * 0.5 + director_sim * 0.3 + year_sim * 0.2
            similarity = max(similarity, sim)
        
        base_score = movie["rating"]
        return base_score * (0.5 + similarity * 0.5)
    
    def recommend(self, user_id, movies, user_ratings, top_n=5):
        """为用户推荐Top-N电影"""
        # 已看过的电影
        watched = set(user_ratings.keys())
        
        # 候选电影
        candidates = [m for m in movies if m["id"] not in watched]
        
        # 计算每部电影的预测评分
        scores = []
        for movie in candidates:
            score = self.predict_rating(user_id, movie, user_ratings)
            scores.append((movie, score))
        
        # 排序并返回Top-N
        scores.sort(key=lambda x: x[1], reverse=True)
        recommendations = [{"movie": m, "score": s} for m, s in scores[:top_n]]
        
        return recommendations


# ==================== 您提供的爬虫类（完全保留） ====================
class DoubanSpider:
    """豆瓣电影评论爬虫 - 使用您提供的代码"""
    
    def __init__(self):
        self.headers = {
            'Cookie': 'cookiell="118254"; bid=BwcycY_GcAA; _pk_id.100001.4cf6=3a9e7e0cc60384ca.1774349925.; _vwo_uuid_v2=DE05BC0B267794C027A0B55F121E447FB|67213e974038f3093b609c7cfb2d4998; __yadk_uid=WlOmJpAAxxz8cTGpBqTGzHf1dBFSvk6k; __utma=30149280.1808011285.1774349897.1774352743.1774582003.3; __utmc=30149280; __utmz=30149280.1774582003.3.3.utmcsr=cn.bing.com|utmccn=(referral)|utmcmd=referral|utmcct=/; __utma=223695111.406837352.1774349924.1774352743.1774582006.3; __utmb=223695111.0.10.1774582006; __utmc=223695111; __utmz=223695111.1774582006.3.3.utmcsr=cn.bing.com|utmccn=(referral)|utmcmd=referral|utmcct=/; _pk_ref.100001.4cf6=%5B%22%22%2C%22%22%2C1774582007%2C%22https%3A%2F%2Fcn.bing.com%2F%22%5D; _pk_ses.100001.4cf6=1; ap_v=0,6.0; dbcl2="293423119:3LWQOX7tv8w"; ck=Fula; push_noty_num=0; push_doumail_num=0; frodotk_db="f1797be7334cb148d46c73b599e613d8"; __utmv=30149280.29342; __utmb=30149280.5.10.1774582003',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Encoding': 'gzip, deflate',
            'Host': 'movie.douban.com',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.4 Safari/605.1.15',
            'Accept-Language': 'zh-CN,zh-Hans;q=0.9',
            'Referer': 'https://movie.douban.com/subject/35267224/?from=showing',
            'Connection': 'keep-alive'
        }
        self.ml_analyzer = None
        self.save_dir = r"D:\2026实训python\爬虫Data"
        
        if not os.path.exists(self.save_dir):
            os.makedirs(self.save_dir)
    
    def search_movie_id(self, movie_name):
        try:
            search_url = f"https://movie.douban.com/j/subject_suggest?q={movie_name}"
            response = requests.get(search_url, headers=self.headers, timeout=10)
            
            if response.status_code == 200:
                results = response.json()
                if results:
                    return results[0].get('id'), results[0].get('title')
            return None, None
        except Exception as e:
            print(f"豆瓣搜索失败: {e}")
            return None, None
    
    def get_movie_detail(self, movie_id):
        try:
            url = f"https://movie.douban.com/subject/{movie_id}/"
            response = requests.get(url, headers=self.headers, timeout=10)
            response.encoding = 'utf-8'
            
            if response.status_code != 200:
                return None
            
            soup = BeautifulSoup(response.text, 'html.parser')
            info = {}
            
            director_elem = soup.find('a', rel='v:directedBy')
            info['director'] = director_elem.text.strip() if director_elem else "未知"
            
            actor_elems = soup.find_all('a', rel='v:starring')
            info['actors'] = [a.text.strip() for a in actor_elems[:5]] if actor_elems else ["未知"]
            
            genre_elems = soup.find_all('span', property='v:genre')
            info['genres'] = [g.text.strip() for g in genre_elems] if genre_elems else ["未知"]
            
            info_elem = soup.find('div', id='info')
            if info_elem:
                info_text = info_elem.text
                country_match = re.search(r'制片国家/地区:\s*(.+?)(?:\n|$)', info_text)
                info['country'] = country_match.group(1).strip() if country_match else "未知"
                language_match = re.search(r'语言:\s*(.+?)(?:\n|$)', info_text)
                info['language'] = language_match.group(1).strip() if language_match else "未知"
                duration_match = re.search(r'片长:\s*(.+?)(?:\n|$)', info_text)
                info['duration'] = duration_match.group(1).strip() if duration_match else "未知"
            else:
                info['country'] = "未知"
                info['language'] = "未知"
                info['duration'] = "未知"
            
            rating_elem = soup.find('strong', property='v:average')
            info['rating'] = rating_elem.text.strip() if rating_elem else "暂无"
            
            votes_elem = soup.find('span', property='v:votes')
            info['votes'] = votes_elem.text.strip() if votes_elem else "0"
            
            summary_elem = soup.find('span', property='v:summary')
            info['summary'] = summary_elem.text.strip() if summary_elem else "暂无简介"
            
            return info
            
        except Exception as e:
            print(f"获取电影详情失败: {e}")
            return None
    
    def crawl_reviews(self, movie_id, max_count=20, progress_callback=None):
        reviews = []
        max_page = (max_count + 19) // 20
        max_page = min(max_page, 5)
        
        result_file = os.path.join(self.save_dir, f"douban_comments_{movie_id}.csv")
        
        if os.path.exists(result_file):
            os.remove(result_file)
        
        header = True
        
        for page in range(1, max_page + 1):
            if progress_callback:
                progress_callback(len(reviews), max_count)
            
            url = f'https://movie.douban.com/subject/{movie_id}/comments?start={(page - 1) * 20}&limit=20&status=P&sort=new_score'
            
            try:
                response = requests.get(url, headers=self.headers, verify=False, timeout=15)
                soup = BeautifulSoup(response.text, 'html.parser')
                review_items = soup.find_all('div', class_='comment-item')
                
                if not review_items:
                    break
                
                user_name_list = []
                star_list = []
                time_list = []
                ip_list = []
                vote_list = []
                content_list = []
                
                for item in review_items:
                    if len(reviews) >= max_count:
                        break
                    
                    try:
                        user_name = item.find('div', class_='avatar').a.get('title', '')
                    except:
                        user_name = ''
                    user_name_list.append(user_name)
                    
                    try:
                        star_class = item.find('span', class_='rating').get('class', [])[0]
                        star = int(star_class.replace('allstar', '')) / 10
                    except:
                        star = 0
                    star_list.append(star)
                    
                    try:
                        comment_time = item.find('span', class_='comment-time').get('title', '')
                    except:
                        comment_time = ''
                    time_list.append(comment_time)
                    
                    try:
                        ip = item.find('span', class_='comment-location').text
                    except:
                        ip = ''
                    ip_list.append(ip)
                    
                    try:
                        vote = item.find('span', class_='votes').text
                    except:
                        vote = 0
                    vote_list.append(vote)
                    
                    try:
                        content = item.find('span', class_='short').text
                        content = content.replace(',', '，').replace(' ', '').replace('\n', '').replace('\t', '').replace('\r', '')
                    except:
                        content = ''
                    content_list.append(content)
                    
                    reviews.append({
                        "content": content,
                        "rating": int(star),
                        "user": user_name,
                        "time": comment_time,
                        "ip": ip,
                        "votes": vote
                    })
                    
                    if progress_callback:
                        progress_callback(len(reviews), max_count)
                
                df = pd.DataFrame({
                    '页码': page,
                    '评论者昵称': user_name_list,
                    '评论星级': star_list,
                    '评论时间': time_list,
                    '评论者IP属地': ip_list,
                    '有用数': vote_list,
                    '评论内容': content_list,
                })
                
                df.to_csv(result_file, mode='a+', header=header, index=False, encoding='utf_8_sig')
                header = False
                
                time.sleep(random.uniform(1, 3))
                
            except Exception as e:
                print(f"爬取第{page}页失败: {e}")
                break
        
        return reviews


# ==================== PyTorch CNN+LSTM 情感分析模型 ====================

class CNNLSTM(nn.Module):
    """CNN + LSTM 情感分析模型"""
    
    def __init__(self, vocab_size, embedding_dim=128, hidden_dim=128, num_classes=3, dropout=0.5):
        super(CNNLSTM, self).__init__()
        
        self.embedding = nn.Embedding(vocab_size, embedding_dim, padding_idx=0)
        
        self.conv3 = nn.Conv1d(embedding_dim, 64, kernel_size=3, padding=1)
        self.conv4 = nn.Conv1d(embedding_dim, 64, kernel_size=4, padding=1)
        self.conv5 = nn.Conv1d(embedding_dim, 64, kernel_size=5, padding=2)
        
        self.pool = nn.AdaptiveMaxPool1d(1)
        
        self.lstm = nn.LSTM(
            input_size=192,
            hidden_size=hidden_dim,
            num_layers=2,
            batch_first=True,
            dropout=dropout,
            bidirectional=True
        )
        
        self.fc1 = nn.Linear(hidden_dim * 2, 64)
        self.dropout = nn.Dropout(dropout)
        self.fc2 = nn.Linear(64, num_classes)
        self.relu = nn.ReLU()
        
    def forward(self, x):
        embedded = self.embedding(x)
        embedded_cnn = embedded.permute(0, 2, 1)
        
        conv3_out = self.relu(self.conv3(embedded_cnn))
        conv4_out = self.relu(self.conv4(embedded_cnn))
        conv5_out = self.relu(self.conv5(embedded_cnn))
        
        pool3 = self.pool(conv3_out).squeeze(-1)
        pool4 = self.pool(conv4_out).squeeze(-1)
        pool5 = self.pool(conv5_out).squeeze(-1)
        
        cnn_features = torch.cat([pool3, pool4, pool5], dim=1)
        
        lstm_out, _ = self.lstm(embedded)
        lstm_features = lstm_out[:, -1, :]
        
        combined = torch.cat([cnn_features, lstm_features], dim=1)
        
        out = self.relu(self.fc1(combined))
        out = self.dropout(out)
        out = self.fc2(out)
        
        return out


class SentimentDataset(Dataset):
    def __init__(self, texts, labels, word2idx, max_len=100):
        self.texts = texts
        self.labels = labels
        self.word2idx = word2idx
        self.max_len = max_len
        
    def __len__(self):
        return len(self.texts)
    
    def __getitem__(self, idx):
        text = self.texts[idx]
        label = self.labels[idx]
        
        indices = [self.word2idx.get(word, self.word2idx.get('<UNK>', 1)) for word in text.split()]
        
        if len(indices) > self.max_len:
            indices = indices[:self.max_len]
        else:
            indices = indices + [0] * (self.max_len - len(indices))
        
        return torch.tensor(indices, dtype=torch.long), torch.tensor(label, dtype=torch.long)


class PyTorchSentimentAnalyzer:
    def __init__(self, vocab_size=20000, embedding_dim=128, hidden_dim=128, max_len=100):
        self.vocab_size = vocab_size
        self.embedding_dim = embedding_dim
        self.hidden_dim = hidden_dim
        self.max_len = max_len
        self.word2idx = {'<PAD>': 0, '<UNK>': 1}
        self.idx2word = {0: '<PAD>', 1: '<UNK>'}
        self.model = None
        self.is_trained = False
        self.device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
        self.label_map = {'positive': 0, 'neutral': 1, 'negative': 2}
        self.reverse_label_map = {0: 'positive', 1: 'neutral', 2: 'negative'}
        
    def preprocess_text(self, text):
        text = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9]', ' ', text)
        words = jieba.cut(text)
        stopwords = {'的', '了', '是', '在', '我', '有', '和', '就', '不', '人', '都'}
        words = [w for w in words if w not in stopwords and len(w.strip()) > 0]
        return ' '.join(words)
    
    def build_vocab(self, texts):
        word_count = {}
        for text in texts:
            for word in text.split():
                word_count[word] = word_count.get(word, 0) + 1
        
        sorted_words = sorted(word_count.items(), key=lambda x: x[1], reverse=True)
        
        for i, (word, _) in enumerate(sorted_words[:self.vocab_size - 2]):
            self.word2idx[word] = len(self.word2idx)
            self.idx2word[len(self.idx2word)] = word
        
        print(f"词表大小: {len(self.word2idx)}")
    
    def prepare_data(self, texts, labels, test_size=0.2, val_size=0.1):
        from sklearn.model_selection import train_test_split
        
        processed_texts = [self.preprocess_text(t) for t in texts]
        
        self.build_vocab(processed_texts)
        
        encoded_labels = [self.label_map[l] for l in labels]
        
        X_temp, X_test, y_temp, y_test = train_test_split(
            processed_texts, encoded_labels, test_size=test_size, random_state=42, stratify=encoded_labels
        )
        
        X_train, X_val, y_train, y_val = train_test_split(
            X_temp, y_temp, test_size=val_size/(1-test_size), random_state=42, stratify=y_temp
        )
        
        train_dataset = SentimentDataset(X_train, y_train, self.word2idx, self.max_len)
        val_dataset = SentimentDataset(X_val, y_val, self.word2idx, self.max_len)
        test_dataset = SentimentDataset(X_test, y_test, self.word2idx, self.max_len)
        
        print(f"训练集: {len(train_dataset)} 条")
        print(f"验证集: {len(val_dataset)} 条")
        print(f"测试集: {len(test_dataset)} 条")
        
        return train_dataset, val_dataset, test_dataset
    
    def build_model(self):
        self.model = CNNLSTM(
            vocab_size=len(self.word2idx),
            embedding_dim=self.embedding_dim,
            hidden_dim=self.hidden_dim,
            num_classes=3
        ).to(self.device)
        return self.model
    
    def train(self, train_dataset, val_dataset, epochs=15, batch_size=32, lr=0.001):
        train_loader = DataLoader(train_dataset, batch_size=batch_size, shuffle=True)
        val_loader = DataLoader(val_dataset, batch_size=batch_size, shuffle=False)
        
        optimizer = optim.Adam(self.model.parameters(), lr=lr)
        criterion = nn.CrossEntropyLoss()
        scheduler = optim.lr_scheduler.ReduceLROnPlateau(optimizer, mode='min', factor=0.5, patience=3)
        
        print("开始训练情感分析模型...")
        for epoch in range(epochs):
            self.model.train()
            total_loss = 0
            correct = 0
            total = 0
            
            for data, target in train_loader:
                data, target = data.to(self.device), target.to(self.device)
                
                optimizer.zero_grad()
                output = self.model(data)
                loss = criterion(output, target)
                loss.backward()
                optimizer.step()
                
                total_loss += loss.item()
                _, predicted = torch.max(output.data, 1)
                total += target.size(0)
                correct += (predicted == target).sum().item()
            
            train_loss = total_loss / len(train_loader)
            train_acc = 100 * correct / total
            
            self.model.eval()
            val_loss = 0
            correct = 0
            total = 0
            
            with torch.no_grad():
                for data, target in val_loader:
                    data, target = data.to(self.device), target.to(self.device)
                    output = self.model(data)
                    loss = criterion(output, target)
                    
                    val_loss += loss.item()
                    _, predicted = torch.max(output.data, 1)
                    total += target.size(0)
                    correct += (predicted == target).sum().item()
            
            val_loss = val_loss / len(val_loader)
            val_acc = 100 * correct / total
            
            scheduler.step(val_loss)
            
            print(f'Epoch {epoch+1}/{epochs} - Train Loss: {train_loss:.4f}, Train Acc: {train_acc:.2f}%, Val Loss: {val_loss:.4f}, Val Acc: {val_acc:.2f}%')
        
        self.is_trained = True
    
    def predict_sentiment(self, text):
        if not self.is_trained or self.model is None:
            return "neutral", 0.5
        
        processed = self.preprocess_text(text)
        indices = [self.word2idx.get(word, self.word2idx.get('<UNK>', 1)) for word in processed.split()]
        
        if len(indices) > self.max_len:
            indices = indices[:self.max_len]
        else:
            indices = indices + [0] * (self.max_len - len(indices))
        
        data = torch.tensor([indices], dtype=torch.long).to(self.device)
        
        self.model.eval()
        with torch.no_grad():
            output = self.model(data)
            probs = torch.softmax(output, dim=1)
            _, predicted = torch.max(output, 1)
        
        sentiment = self.reverse_label_map[predicted.item()]
        confidence = float(probs[0][predicted.item()].item())
        
        return sentiment, confidence
    
    def predict_batch(self, texts):
        results = []
        for text in texts:
            sentiment, confidence = self.predict_sentiment(text)
            results.append({
                'sentiment': sentiment,
                'confidence': confidence
            })
        return results
    
    def save_model(self, model_path='pytorch_model.pth', vocab_path='vocab.pkl'):
        torch.save(self.model.state_dict(), model_path)
        with open(vocab_path, 'wb') as f:
            pickle.dump({
                'word2idx': self.word2idx,
                'idx2word': self.idx2word,
                'max_len': self.max_len
            }, f)
    
    def load_model(self, model_path='pytorch_model.pth', vocab_path='vocab.pkl'):
        self.build_model()
        self.model.load_state_dict(torch.load(model_path, map_location=self.device))
        with open(vocab_path, 'rb') as f:
            vocab = pickle.load(f)
            self.word2idx = vocab['word2idx']
            self.idx2word = vocab['idx2word']
            self.max_len = vocab['max_len']
        self.is_trained = True


# ==================== 数据管理类 ====================
class MovieDataManager:
    """电影数据管理器"""
    
    def __init__(self):
        self.movies = self._generate_movies()
        self.reviews = self._generate_reviews()
        self.users = self._generate_users()
        self.current_user = "user_001"
        self.crawled_reviews = {}
        self.uploaded_reviews = {}
        self.sentiment_analyzer = PyTorchSentimentAnalyzer()
        self.recommender = HybridRecommender()
        self.spider = DoubanSpider() if REQUESTS_AVAILABLE else None
        self.recommender_trained = False
        
        if self.spider:
            self.spider.ml_analyzer = self.sentiment_analyzer
    
    def _generate_movies(self):
        return [
            {"id": 1, "title": "肖申克的救赎", "director": "弗兰克·德拉邦特", "actors": ["蒂姆·罗宾斯", "摩根·弗里曼"], 
             "genre": "剧情", "year": 1994, "rating": 9.7, "description": "希望让人自由", "duration": 142, "country": "美国", "language": "英语"},
            {"id": 2, "title": "霸王别姬", "director": "陈凯歌", "actors": ["张国荣", "张丰毅", "巩俐"], 
             "genre": "剧情", "year": 1993, "rating": 9.6, "description": "风华绝代", "duration": 171, "country": "中国", "language": "汉语普通话"},
            {"id": 3, "title": "阿甘正传", "director": "罗伯特·泽米吉斯", "actors": ["汤姆·汉克斯"], 
             "genre": "剧情", "year": 1994, "rating": 9.5, "description": "人生就像巧克力", "duration": 142, "country": "美国", "language": "英语"},
            {"id": 4, "title": "这个杀手不太冷", "director": "吕克·贝松", "actors": ["让·雷诺", "娜塔莉·波特曼"], 
             "genre": "剧情/动作", "year": 1994, "rating": 9.4, "description": "杀手与小女孩", "duration": 110, "country": "法国", "language": "英语"},
            {"id": 5, "title": "泰坦尼克号", "director": "詹姆斯·卡梅隆", "actors": ["莱昂纳多", "凯特·温丝莱特"], 
             "genre": "爱情/灾难", "year": 1997, "rating": 9.5, "description": "永恒的爱情", "duration": 194, "country": "美国", "language": "英语"},
            {"id": 6, "title": "盗梦空间", "director": "克里斯托弗·诺兰", "actors": ["莱昂纳多"], 
             "genre": "科幻/悬疑", "year": 2010, "rating": 9.4, "description": "梦境与现实", "duration": 148, "country": "美国", "language": "英语"},
            {"id": 7, "title": "楚门的世界", "director": "彼得·威尔", "actors": ["金·凯瑞"], 
             "genre": "剧情/科幻", "year": 1998, "rating": 9.3, "description": "真实与虚假", "duration": 103, "country": "美国", "language": "英语"},
            {"id": 8, "title": "千与千寻", "director": "宫崎骏", "actors": ["柊瑠美"], 
             "genre": "动画/奇幻", "year": 2001, "rating": 9.4, "description": "成长的旅程", "duration": 125, "country": "日本", "language": "日语"},
            {"id": 9, "title": "星际穿越", "director": "克里斯托弗·诺兰", "actors": ["马修·麦康纳"], 
             "genre": "科幻", "year": 2014, "rating": 9.4, "description": "穿越时空的爱", "duration": 169, "country": "美国", "language": "英语"},
            {"id": 10, "title": "海上钢琴师", "director": "朱塞佩·托纳多雷", "actors": ["蒂姆·罗斯"], 
             "genre": "剧情/音乐", "year": 1998, "rating": 9.3, "description": "1900的故事", "duration": 125, "country": "意大利", "language": "英语"},
        ]
    
    def _generate_reviews(self):
        reviews = []
        for movie in self.movies:
            for _ in range(random.randint(10, 15)):
                sentiment = random.choice(["positive", "neutral", "negative"])
                text = random.choice(DOUBAN_REAL_REVIEWS[sentiment])
                rating = 5 if sentiment == "positive" else (3 if sentiment == "neutral" else 2)
                
                reviews.append({
                    "movie_id": movie["id"],
                    "movie_name": movie["title"],
                    "user": random.choice(USER_NAMES) + str(random.randint(100, 999)),
                    "content": text,
                    "rating": rating,
                    "sentiment": sentiment,
                    "time": datetime.now().strftime("%Y-%m-%d")
                })
        return reviews
    
    def _generate_users(self):
        users = {}
        for i in range(1, 21):
            user_id = f"user_{i:03d}"
            watched = random.sample(self.movies, random.randint(3, 6))
            users[user_id] = {
                "watched": [m["id"] for m in watched],
                "ratings": {m["id"]: random.randint(6, 10) for m in watched},
                "favorites": [],
                "watchlist": []
            }
        return users
    
    def get_system_reviews(self, movie_id):
        movie = next((m for m in self.movies if m["id"] == movie_id), None)
        if not movie:
            return []
        return [r for r in self.reviews if r["movie_id"] == movie_id]
    
    def get_crawled_reviews(self, movie_name):
        return self.crawled_reviews.get(movie_name, [])
    
    def get_uploaded_reviews(self, movie_name):
        return self.uploaded_reviews.get(movie_name, [])
    
    def crawl_movie_reviews(self, movie_name, max_count, progress_callback=None):
        if not self.spider:
            return None, "爬虫库未安装"
        
        try:
            movie_id, full_title = self.spider.search_movie_id(movie_name)
            if not movie_id:
                return None, f"未找到电影: {movie_name}"
            
            self.spider.ml_analyzer = self.sentiment_analyzer
            
            reviews = self.spider.crawl_reviews(movie_id, max_count, progress_callback)
            
            if not reviews:
                return None, f"未爬取到评论"
            
            self.crawled_reviews[full_title] = reviews
            return reviews, full_title
            
        except Exception as e:
            return None, f"爬取失败: {str(e)}"
    
    def train_recommender(self):
        """训练推荐模型"""
        # 准备用户交互数据
        user_interactions = {}
        for user_id, user_data in self.users.items():
            interactions = {}
            for movie_id, rating in user_data["ratings"].items():
                interactions[movie_id] = rating
            if interactions:
                user_interactions[user_id] = interactions
        
        # 准备电影列表
        all_movies = self.movies.copy()
        for movie in LOCAL_MOVIE_CACHE.values():
            if not any(m["id"] == int(movie["id"]) for m in all_movies):
                all_movies.append({
                    "id": int(movie["id"]),
                    "title": movie["title"],
                    "director": movie["director"],
                    "actors": movie["actors"],
                    "genre": movie["genre"],
                    "year": int(movie["year"]),
                    "rating": float(movie["rating"]),
                    "description": movie["description"],
                    "duration": movie["duration"],
                    "country": movie["country"],
                    "language": movie["language"]
                })
        
        # 训练
        success = self.recommender.train(user_interactions, all_movies, epochs=30)
        if success:
            self.recommender_trained = True
        return success
    
    def get_recommendations(self, user_id, top_n=5):
        """获取推荐列表"""
        user = self.users.get(user_id, {"ratings": {}, "watched": []})
        
        # 准备所有电影
        all_movies = []
        for movie in self.movies:
            all_movies.append({
                "id": movie["id"],
                "title": movie["title"],
                "director": movie["director"],
                "actors": movie["actors"],
                "genre": movie["genre"],
                "year": movie["year"],
                "rating": movie["rating"],
                "description": movie["description"]
            })
        
        # 添加本地缓存电影
        for movie in LOCAL_MOVIE_CACHE.values():
            if not any(m["id"] == int(movie["id"]) for m in all_movies):
                all_movies.append({
                    "id": int(movie["id"]),
                    "title": movie["title"],
                    "director": movie["director"],
                    "actors": movie["actors"],
                    "genre": movie["genre"],
                    "year": int(movie["year"]),
                    "rating": float(movie["rating"]),
                    "description": movie["description"]
                })
        
        # 获取推荐
        recommendations = self.recommender.recommend(user_id, all_movies, user["ratings"], top_n)
        return recommendations
    
    def search_movie_info(self, movie_name):
        result = {"success": False, "data": None, "message": "", "suggestions": []}
        
        if movie_name in LOCAL_MOVIE_CACHE:
            cached = LOCAL_MOVIE_CACHE[movie_name]
            result["success"] = True
            result["data"] = {
                "title": cached["title"],
                "year": cached["year"],
                "rating": cached["rating"],
                "director": cached["director"],
                "actors": cached["actors"],
                "genre": cached["genre"],
                "country": cached.get("country", "未知"),
                "language": cached.get("language", "未知"),
                "duration": cached.get("duration", "未知"),
                "description": cached.get("description", f"经典电影《{cached['title']}》，豆瓣评分{cached['rating']}分"),
                "source": "本地缓存"
            }
            result["message"] = f"从本地缓存找到电影: {movie_name}"
            return result
        
        best_match, score = fuzzy_match_movie(movie_name)
        if best_match and score > 0.6:
            cached = LOCAL_MOVIE_CACHE[best_match]
            result["success"] = True
            result["data"] = {
                "title": cached["title"],
                "year": cached["year"],
                "rating": cached["rating"],
                "director": cached["director"],
                "actors": cached["actors"],
                "genre": cached["genre"],
                "country": cached.get("country", "未知"),
                "language": cached.get("language", "未知"),
                "duration": cached.get("duration", "未知"),
                "description": cached.get("description", f"经典电影《{cached['title']}》，豆瓣评分{cached['rating']}分"),
                "source": f"本地缓存(模糊匹配，相似度:{score:.0%})"
            }
            result["message"] = f"通过模糊匹配找到: {best_match} (相似度: {score:.0%})"
            return result
        
        if self.spider:
            try:
                movie_id, title = self.spider.search_movie_id(movie_name)
                if movie_id:
                    detail = self.spider.get_movie_detail(movie_id)
                    if detail:
                        result["success"] = True
                        result["data"] = {
                            "title": title,
                            "year": detail.get("year", "未知"),
                            "rating": detail.get("rating", "暂无"),
                            "director": detail.get("director", "未知"),
                            "actors": detail.get("actors", ["未知"]),
                            "genre": "/".join(detail.get("genres", ["未知"])),
                            "country": detail.get("country", "未知"),
                            "language": detail.get("language", "未知"),
                            "duration": detail.get("duration", "未知"),
                            "description": detail.get("summary", "暂无简介"),
                            "votes": detail.get("votes", "0"),
                            "source": "豆瓣搜索"
                        }
                        result["message"] = f"从豆瓣找到电影: {title}"
                        return result
            except Exception as e:
                print(f"豆瓣搜索异常: {e}")
        
        suggestions = []
        for cached_name in LOCAL_MOVIE_CACHE.keys():
            if movie_name.lower() in cached_name.lower() or cached_name.lower() in movie_name.lower():
                suggestions.append(cached_name)
        
        if suggestions:
            result["suggestions"] = suggestions[:5]
            result["message"] = f"未找到《{movie_name}》，您是不是想找: {', '.join(suggestions[:3])}"
        else:
            result["message"] = f"未找到电影《{movie_name}》，请尝试使用完整电影名称或选择内置电影"
        
        return result
    
    def parse_uploaded_file(self, file_path, movie_name):
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext == '.json':
            return self._parse_json_file(file_path, movie_name)
        elif file_ext == '.csv':
            return self._parse_csv_file(file_path, movie_name)
        elif file_ext == '.txt':
            return self._parse_txt_file(file_path, movie_name)
        elif file_ext in ['.xlsx', '.xls']:
            try:
                import openpyxl
                return self._parse_excel_file(file_path, movie_name)
            except:
                raise Exception("Excel支持未安装，请运行: pip install openpyxl")
        else:
            raise Exception(f"不支持的文件格式: {file_ext}")
    
    def _parse_json_file(self, file_path, movie_name):
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        reviews = []
        if isinstance(data, list):
            for item in data:
                content = item.get('评论内容', item.get('content', ''))
                user = item.get('用户', item.get('user', ''))
                rating = item.get('评分', item.get('rating', 3))
                if content:
                    reviews.append({
                        "content": content,
                        "user": user,
                        "rating": rating
                    })
        elif isinstance(data, dict):
            for key in data:
                if '评论' in key or 'review' in key.lower():
                    items = data[key]
                    if isinstance(items, list):
                        for item in items:
                            if isinstance(item, dict):
                                content = item.get('评论内容', item.get('content', ''))
                                user = item.get('用户', item.get('user', ''))
                                rating = item.get('评分', item.get('rating', 3))
                                if content:
                                    reviews.append({
                                        "content": content,
                                        "user": user,
                                        "rating": rating
                                    })
                            elif isinstance(item, str):
                                reviews.append({
                                    "content": item,
                                    "user": "",
                                    "rating": 3
                                })
        
        if not reviews:
            raise Exception("JSON文件中没有找到有效的评论数据")
        
        return reviews, movie_name
    
    def _parse_csv_file(self, file_path, movie_name):
        df = pd.read_csv(file_path, encoding='utf-8')
        
        reviews = []
        for _, row in df.iterrows():
            content = None
            for col in ['评论内容', 'content', '评论', 'text']:
                if col in df.columns:
                    content = row[col]
                    break
            
            if not content or pd.isna(content):
                continue
            
            user = None
            for col in ['用户', 'user', 'username']:
                if col in df.columns:
                    user = row[col]
                    break
            
            rating = 3
            for col in ['评分', 'rating', 'score']:
                if col in df.columns and not pd.isna(row[col]):
                    try:
                        rating = int(float(row[col]))
                        rating = max(1, min(5, rating))
                    except:
                        pass
                    break
            
            reviews.append({
                "content": str(content),
                "user": str(user) if user else "",
                "rating": rating
            })
        
        if not reviews:
            raise Exception("CSV文件中没有找到有效的评论数据")
        
        return reviews, movie_name
    
    def _parse_txt_file(self, file_path, movie_name):
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = [line.strip() for line in f.readlines() if line.strip()]
        
        if not lines:
            raise Exception("文件为空")
        
        reviews = []
        for line in lines:
            if '|' in line:
                parts = line.split('|')
                if len(parts) >= 2:
                    user = parts[0].strip()
                    content = parts[1].strip()
                    rating = int(parts[2].strip()) if len(parts) >= 3 else 3
                    reviews.append({
                        "content": content,
                        "user": user,
                        "rating": rating
                    })
                else:
                    reviews.append({
                        "content": line,
                        "user": "",
                        "rating": 3
                    })
            else:
                reviews.append({
                    "content": line,
                    "user": "",
                    "rating": 3
                })
        
        return reviews, movie_name
    
    def _parse_excel_file(self, file_path, movie_name):
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        headers = []
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=1, column=col).value
            if val:
                headers.append(str(val).strip())
        
        content_col = None
        user_col = None
        rating_col = None
        
        for i, header in enumerate(headers):
            header_lower = header.lower()
            if header_lower in ['评论内容', 'content', '评论', 'text']:
                content_col = i
            if header_lower in ['用户', 'user', 'username']:
                user_col = i
            if header_lower in ['评分', 'rating', 'score']:
                rating_col = i
        
        if content_col is None:
            raise Exception("Excel文件必须包含评论内容列")
        
        reviews = []
        for row in range(2, sheet.max_row + 1):
            content = sheet.cell(row=row, column=content_col + 1).value
            if not content:
                continue
            
            content = str(content).strip()
            
            user = ""
            if user_col is not None:
                user_val = sheet.cell(row=row, column=user_col + 1).value
                if user_val:
                    user = str(user_val)
            
            rating = 3
            if rating_col is not None:
                rating_val = sheet.cell(row=row, column=rating_col + 1).value
                if rating_val:
                    try:
                        rating = int(float(rating_val))
                        rating = max(1, min(5, rating))
                    except:
                        pass
            
            reviews.append({
                "content": content,
                "user": user,
                "rating": rating
            })
        
        if not reviews:
            raise Exception("Excel文件中没有找到有效的评论数据")
        
        return reviews, movie_name
    
    def upload_reviews(self, file, movie_name):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        try:
            reviews, movie_name = self.parse_uploaded_file(file_path, movie_name)
            
            texts = [r['content'] for r in reviews]
            results = self.sentiment_analyzer.predict_batch(texts)
            
            for i, review in enumerate(reviews):
                review['sentiment'] = results[i]['sentiment']
                review['confidence'] = results[i]['confidence']
                review['time'] = datetime.now().strftime("%Y-%m-%d %H:%M")
            
            self.uploaded_reviews[movie_name] = reviews
            return reviews, movie_name
            
        finally:
            if os.path.exists(file_path):
                os.remove(file_path)
    
    def train_sentiment_model(self):
        texts = []
        labels = []
        
        for sentiment, reviews in DOUBAN_REAL_REVIEWS.items():
            for review in reviews:
                texts.append(review)
                labels.append(sentiment)
        
        train_dataset, val_dataset, test_dataset = self.sentiment_analyzer.prepare_data(texts, labels)
        
        self.sentiment_analyzer.build_model()
        self.sentiment_analyzer.train(train_dataset, val_dataset, epochs=15)
        
        self.sentiment_analyzer.save_model()
        
        return True
    
    def get_all_data_sources(self):
        sources = {
            "system": [{"name": m["title"], "type": "system", "count": len(self.get_system_reviews(m["id"]))} for m in self.movies],
            "crawled": [{"name": name, "type": "crawled", "count": len(reviews)} for name, reviews in self.crawled_reviews.items()],
            "uploaded": [{"name": name, "type": "uploaded", "count": len(reviews)} for name, reviews in self.uploaded_reviews.items()]
        }
        return sources


# 创建全局数据管理器
data_manager = MovieDataManager()

# 尝试加载已有模型
try:
    data_manager.sentiment_analyzer.load_model()
    print("✅ PyTorch情感分析模型加载成功")
except:
    print("⚠️ 未找到情感分析模型，请先训练模型")

# 尝试训练推荐模型
try:
    data_manager.train_recommender()
    print("✅ 神经网络推荐模型训练完成")
except Exception as e:
    print(f"⚠️ 推荐模型训练失败: {e}，将使用降级推荐")

# ==================== 路由定义 ====================

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/data_collection')
def data_collection():
    movies = data_manager.movies
    return render_template('data_collection.html', movies=movies)

@app.route('/data_analysis')
def data_analysis():
    sources = data_manager.get_all_data_sources()
    return render_template('data_analysis.html', sources=sources)

@app.route('/movie_analysis')
def movie_analysis():
    movies = data_manager.movies
    return render_template('movie_analysis.html', movies=movies)

@app.route('/recommendation')
def recommendation():
    return render_template('recommendation.html')

@app.route('/ai_assistant')
def ai_assistant():
    return render_template('ai_assistant.html')

@app.route('/social')
def social():
    return render_template('social.html')

@app.route('/user')
def user():
    return render_template('user.html', user=data_manager.current_user)

# ==================== API接口 ====================

@app.route('/api/get_data_sources', methods=['GET'])
def api_get_data_sources():
    sources = data_manager.get_all_data_sources()
    return jsonify({'success': True, 'sources': sources})

@app.route('/api/crawl_reviews', methods=['POST'])
def api_crawl_reviews():
    data = request.json
    movie_name = data.get('movie_name')
    max_count = int(data.get('max_count', 20))
    
    def progress_callback(current, total):
        pass
    
    reviews, result = data_manager.crawl_movie_reviews(movie_name, max_count, progress_callback)
    
    if reviews:
        return jsonify({
            'success': True,
            'reviews': reviews,
            'movie_name': result,
            'count': len(reviews)
        })
    else:
        return jsonify({
            'success': False,
            'error': result
        })

@app.route('/api/search_movie', methods=['POST'])
def api_search_movie():
    data = request.json
    movie_name = data.get('movie_name', '').strip()
    
    if not movie_name:
        return jsonify({'success': False, 'error': '请输入电影名称'})
    
    result = data_manager.search_movie_info(movie_name)
    return jsonify(result)

@app.route('/api/recommend', methods=['POST'])
def api_recommend():
    """神经网络推荐接口"""
    data = request.json
    user_id = data.get('user_id', data_manager.current_user)
    top_n = int(data.get('top_n', 5))
    
    try:
        recommendations = data_manager.get_recommendations(user_id, top_n)
        
        # 格式化返回结果
        result = []
        for rec in recommendations:
            movie = rec["movie"]
            result.append({
                "id": movie["id"],
                "title": movie["title"],
                "director": movie["director"],
                "actors": movie["actors"],
                "genre": movie["genre"],
                "year": movie["year"],
                "rating": movie["rating"],
                "description": movie["description"],
                "score": round(rec["score"], 2)
            })
        
        return jsonify({
            'success': True,
            'recommendations': result,
            'method': '神经网络推荐 (协同过滤 + 内容特征)',
            'user_id': user_id
        })
    except Exception as e:
        # 降级推荐
        user = data_manager.users.get(user_id, {"ratings": {}, "watched": []})
        watched = set(user["watched"])
        
        recs = [m for m in data_manager.movies if m["id"] not in watched]
        recs.sort(key=lambda x: x["rating"], reverse=True)
        
        fallback_result = []
        for m in recs[:top_n]:
            fallback_result.append({
                "id": m["id"],
                "title": m["title"],
                "director": m["director"],
                "actors": m["actors"],
                "genre": m["genre"],
                "year": m["year"],
                "rating": m["rating"],
                "description": m["description"],
                "score": m["rating"]
            })
        
        return jsonify({
            'success': True,
            'recommendations': fallback_result,
            'method': '降级推荐 (基于评分)',
            'warning': str(e)
        })

@app.route('/api/train_recommender', methods=['POST'])
def api_train_recommender():
    """训练推荐模型"""
    try:
        success = data_manager.train_recommender()
        if success:
            return jsonify({'success': True, 'message': '推荐模型训练完成！'})
        else:
            return jsonify({'success': False, 'message': '训练数据不足，请先添加更多观影记录'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/upload_reviews', methods=['POST'])
def api_upload_reviews():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '没有文件'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': '文件名为空'})
    
    movie_name = request.form.get('movie_name', os.path.splitext(file.filename)[0])
    
    try:
        reviews, movie_name = data_manager.upload_reviews(file, movie_name)
        return jsonify({
            'success': True,
            'reviews': reviews,
            'movie_name': movie_name,
            'count': len(reviews)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/analyze_sentiment', methods=['POST'])
def api_analyze_sentiment():
    data = request.json
    source_type = data.get('source_type')
    name = data.get('name')
    
    if source_type == 'system':
        movie = next((m for m in data_manager.movies if m["title"] == name), None)
        if movie:
            reviews = data_manager.get_system_reviews(movie["id"])
        else:
            reviews = []
    elif source_type == 'crawled':
        reviews = data_manager.get_crawled_reviews(name)
    elif source_type == 'uploaded':
        reviews = data_manager.get_uploaded_reviews(name)
    else:
        return jsonify({'success': False, 'error': '无效的数据源类型'})
    
    if not reviews:
        return jsonify({'success': False, 'error': '暂无评论数据'})
    
    texts = [r['content'] for r in reviews]
    results = data_manager.sentiment_analyzer.predict_batch(texts)
    
    for i, review in enumerate(reviews):
        review['sentiment'] = results[i]['sentiment']
        review['confidence'] = results[i]['confidence']
    
    sentiments = [r['sentiment'] for r in reviews]
    counts = Counter(sentiments)
    total = len(reviews)
    
    return jsonify({
        'success': True,
        'name': name,
        'source_type': source_type,
        'total': total,
        'positive': counts.get('positive', 0),
        'positive_pct': counts.get('positive', 0) / total * 100 if total > 0 else 0,
        'neutral': counts.get('neutral', 0),
        'neutral_pct': counts.get('neutral', 0) / total * 100 if total > 0 else 0,
        'negative': counts.get('negative', 0),
        'negative_pct': counts.get('negative', 0) / total * 100 if total > 0 else 0,
        'reviews': reviews[:100]
    })

@app.route('/api/get_rating_dist', methods=['POST'])
def api_get_rating_dist():
    data = request.json
    source_type = data.get('source_type')
    name = data.get('name')
    
    if source_type == 'system':
        movie = next((m for m in data_manager.movies if m["title"] == name), None)
        if movie:
            reviews = data_manager.get_system_reviews(movie["id"])
        else:
            reviews = []
    elif source_type == 'crawled':
        reviews = data_manager.get_crawled_reviews(name)
    elif source_type == 'uploaded':
        reviews = data_manager.get_uploaded_reviews(name)
    else:
        return jsonify({'success': False, 'error': '无效的数据源类型'})
    
    if not reviews:
        return jsonify({'success': False, 'error': '暂无评论数据'})
    
    ratings = [r["rating"] for r in reviews]
    counts = [ratings.count(i) for i in range(1, 6)]
    
    fig, ax = plt.subplots(figsize=(8, 5))
    bars = ax.bar(["1星", "2星", "3星", "4星", "5星"], counts, 
                   color=['#ff6b6b', '#ffa07a', '#ffd966', '#98d98e', '#6bcf7f'])
    ax.set_title(f"{name} 评分分布", fontsize=14, fontweight='bold')
    ax.set_xlabel("评分", fontsize=12)
    ax.set_ylabel("评论数量", fontsize=12)
    
    for bar, count in zip(bars, counts):
        if count > 0:
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5, 
                   str(count), ha='center', va='bottom', fontsize=10)
    
    buffer = BytesIO()
    fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
    buffer.seek(0)
    image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    plt.close(fig)
    
    return jsonify({
        'success': True,
        'image': image_base64
    })

@app.route('/api/get_trend_chart', methods=['POST'])
def api_get_trend_chart():
    data = request.json
    source_type = data.get('source_type')
    name = data.get('name')
    
    if source_type == 'system':
        movie = next((m for m in data_manager.movies if m["title"] == name), None)
        if movie:
            reviews = data_manager.get_system_reviews(movie["id"])
        else:
            reviews = []
    elif source_type == 'crawled':
        reviews = data_manager.get_crawled_reviews(name)
    elif source_type == 'uploaded':
        reviews = data_manager.get_uploaded_reviews(name)
    else:
        return jsonify({'success': False, 'error': '无效的数据源类型'})
    
    if len(reviews) < 2:
        return jsonify({'success': False, 'error': '评论数量不足，无法生成趋势图'})
    
    score_map = {"positive": 1, "neutral": 0, "negative": -1}
    scores = [score_map[r["sentiment"]] for r in reviews]
    
    fig, ax = plt.subplots(figsize=(10, 5))
    x = range(len(scores))
    ax.plot(x, scores, 'b-', linewidth=2, marker='o', markersize=3, alpha=0.7)
    ax.axhline(y=1, color='green', linestyle='--', alpha=0.5, label='正面')
    ax.axhline(y=0, color='gray', linestyle='--', alpha=0.5, label='中性')
    ax.axhline(y=-1, color='red', linestyle='--', alpha=0.5, label='负面')
    ax.fill_between(x, scores, 0, where=(np.array(scores) > 0), color='green', alpha=0.3)
    ax.fill_between(x, scores, 0, where=(np.array(scores) < 0), color='red', alpha=0.3)
    
    ax.set_title(f"{name} 情感趋势分析", fontsize=14, fontweight='bold')
    ax.set_xlabel("评论序号（按时间顺序）", fontsize=12)
    ax.set_ylabel("情感得分", fontsize=12)
    ax.set_yticks([-1, -0.5, 0, 0.5, 1])
    ax.set_yticklabels(['负面', '偏负面', '中性', '偏正面', '正面'])
    ax.legend(loc='upper right')
    ax.grid(True, alpha=0.3)
    
    buffer = BytesIO()
    fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
    buffer.seek(0)
    image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    plt.close(fig)
    
    return jsonify({
        'success': True,
        'image': image_base64
    })

@app.route('/api/chat', methods=['POST'])
def api_chat():
    data = request.json
    message = data.get('message', '')
    message_lower = message.lower()
    
    if "推荐" in message_lower:
        response = "推荐《肖申克的救赎》《霸王别姬》《阿甘正传》这些经典电影"
    elif "评分" in message_lower:
        response = "《肖申克的救赎》9.7分，《霸王别姬》9.6分，《阿甘正传》9.5分"
    elif "类型" in message_lower:
        response = "电影类型包括：剧情、科幻、动画、爱情、喜剧、悬疑、动作、奇幻等"
    else:
        response = "我是电影助手！可以问我电影推荐、评分查询、类型介绍等问题"
    
    return jsonify({
        'success': True,
        'response': response
    })

@app.route('/api/user/info')
def api_user_info():
    user = data_manager.users.get(data_manager.current_user, {"watched": []})
    watched_movies = [m for m in data_manager.movies if m["id"] in user["watched"]]
    
    return jsonify({
        'username': data_manager.current_user,
        'watched_count': len(watched_movies),
        'watched_movies': watched_movies
    })

@app.route('/api/user/login', methods=['POST'])
def api_user_login():
    data = request.json
    username = data.get('username')
    
    if username in data_manager.users:
        data_manager.current_user = username
        return jsonify({'success': True, 'username': username})
    else:
        return jsonify({'success': False, 'error': '用户不存在'})

@app.route('/api/user/register', methods=['POST'])
def api_user_register():
    data = request.json
    username = data.get('username')
    
    if username not in data_manager.users:
        data_manager.users[username] = {"watched": [], "ratings": {}, "favorites": [], "watchlist": []}
        data_manager.current_user = username
        return jsonify({'success': True, 'username': username})
    else:
        return jsonify({'success': False, 'error': '用户已存在'})

@app.route('/api/user/add_record', methods=['POST'])
def api_user_add_record():
    data = request.json
    movie_title = data.get('movie_title')
    rating = int(data.get('rating', 8))
    
    movie = next((m for m in data_manager.movies if m["title"] == movie_title), None)
    if not movie:
        return jsonify({'success': False, 'error': '电影不存在'})
    
    user = data_manager.users[data_manager.current_user]
    if movie["id"] not in user["watched"]:
        user["watched"].append(movie["id"])
    user["ratings"][movie["id"]] = rating
    
    return jsonify({'success': True})

@app.route('/api/train_model', methods=['POST'])
def api_train_model():
    """训练情感分析模型"""
    try:
        data_manager.train_sentiment_model()
        return jsonify({
            'success': True,
            'message': '情感分析模型训练完成！'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })

if __name__ == '__main__':
    print("=" * 60)
    print("豆瓣电影数据分析与推荐系统 - PyTorch CNN+LSTM版本")
    print("神经网络推荐引擎: 协同过滤 + 内容特征 + MLP")
    print("访问地址: http://127.0.0.1:5000")
    print("=" * 60)
    
    try:
        import torch
        print(f"PyTorch版本: {torch.__version__}")
        print(f"CUDA可用: {torch.cuda.is_available()}")
    except:
        print("⚠️ PyTorch未安装，请运行: pip install torch")
    
    app.run(debug=True, host='0.0.0.0', port=5000)